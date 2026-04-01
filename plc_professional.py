# =============================================================================
# PLC PROFESSIONAL — EVALUACIÓN COGNITIVA
# Archivo único completo — v3.0
#
# CAMBIOS v3.0 (feedback psicología):
#   - Estímulos vectoriales fieles a la imagen de referencia (cruces con cuadrados)
#   - 6 tipos de distractores basados en posición de cuadrados y línea desplazada
#   - 47 estímulos en MÁXIMO 2 FILAS por línea (layout horizontal amplio)
#   - Pre-prueba con target + 3 distractores etiquetados (como la imagen)
#   - Sin recomendaciones finales (herramienta para terapeuta)
#   - Sin patologización → descripción de rendimiento cognitivo
#   - Resultados con: curvas de normalidad (Gauss), estilos de atención,
#     gráfico de barras, curva de comportamiento por línea
#   - Medición por ítem: timestamp ms por clic → velocidad de reacción
#   - Índice de concentración destacado
#   - Siguiente automático al llegar a 0s
# =============================================================================

import tkinter as tk
from tkinter import ttk, messagebox
import random
import time
import os
import sys
import math
import subprocess
from datetime import datetime
import statistics
import numpy as np
import pandas as pd

try:
    import tensorflow as tf
    from tensorflow.keras.models import load_model
    import joblib
    TF_AVAILABLE = True
except ImportError:
    TF_AVAILABLE = False

# ─── Matplotlib para gráficas en resultados ──────────────────────────────────
try:
    import matplotlib
    matplotlib.use('TkAgg')
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    from matplotlib.figure import Figure
    MPL_AVAILABLE = True
except ImportError:
    MPL_AVAILABLE = False


# =============================================================================
# UTILIDADES
# =============================================================================

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


# =============================================================================
# DEFINICIÓN DE ESTÍMULOS — fieles a la imagen de referencia
#
# Cada estímulo es una cruz. Los cuadrados negros (■) se dibujan en los
# extremos de los brazos. La posición y cantidad de cuadrados + la posición
# de la línea horizontal definen si es TARGET o distractor.
#
# TARGET:
#   Cruz simétrica, línea horizontal centrada, cuadrados en ambos extremos
#   del brazo horizontal  →  ■—+—■
#
# DISTRACTORES (basados en la imagen):
#   D1 — cuadrado solo en extremo derecho del horizontal, línea vertical larga arriba
#   D2 — cuadrado solo en extremo izquierdo, línea vertical asimétrica
#   D3 — cuadrado abajo-izquierda, línea horizontal desplazada hacia abajo
#   D4 — sin cuadrados (cruz desnuda)
#   D5 — cuadrado solo arriba del brazo vertical (no en el horizontal)
#   D6 — cruz con cuadrados en vertical (arriba y abajo), sin cuadrados horizontal
# =============================================================================

STIM_TYPES = {
    # TARGET — cruz simétrica, cuadrados en AMBOS extremos horizontales
    'T':  {'is_target': True,  'sq_left': True,  'sq_right': True,
           'sq_top': False, 'sq_bottom': False, 'h_offset': 0,   'v_asym': 0},

    # ── Distractores tipo A: un solo cuadrado en el brazo horizontal ───────
    'D1': {'is_target': False, 'sq_left': False, 'sq_right': True,
           'sq_top': False, 'sq_bottom': False, 'h_offset': 0,   'v_asym': 8},
    'D2': {'is_target': False, 'sq_left': True,  'sq_right': False,
           'sq_top': False, 'sq_bottom': False, 'h_offset': 0,   'v_asym': -6},
    'D3': {'is_target': False, 'sq_left': True,  'sq_right': False,
           'sq_top': False, 'sq_bottom': False, 'h_offset': 10,  'v_asym': 0},

    # ── Distractores tipo B: sin cuadrados en horizontal ──────────────────
    'D4': {'is_target': False, 'sq_left': False, 'sq_right': False,
           'sq_top': False, 'sq_bottom': False, 'h_offset': 0,   'v_asym': 0},
    'D5': {'is_target': False, 'sq_left': False, 'sq_right': False,
           'sq_top': True,  'sq_bottom': False, 'h_offset': 0,   'v_asym': 0},
    'D6': {'is_target': False, 'sq_left': False, 'sq_right': False,
           'sq_top': True,  'sq_bottom': True,  'h_offset': 0,   'v_asym': 0},

    # ── Distractores tipo C: combinaciones mixtas más difíciles ───────────
    # D7: cuadrado derecho + desplazamiento horizontal abajo (suma trampa)
    'D7': {'is_target': False, 'sq_left': False, 'sq_right': True,
           'sq_top': False, 'sq_bottom': False, 'h_offset': -9,  'v_asym': 0},
    # D8: cuadrado izquierdo + brazo vertical asimétrico hacia arriba
    'D8': {'is_target': False, 'sq_left': True,  'sq_right': False,
           'sq_top': False, 'sq_bottom': False, 'h_offset': 0,   'v_asym': 10},
    # D9: cuadrado abajo del vertical (muy poco común → alta dificultad)
    'D9': {'is_target': False, 'sq_left': False, 'sq_right': False,
           'sq_top': False, 'sq_bottom': True,  'h_offset': 0,   'v_asym': 0},
    # D10: un cuadrado a la derecha + línea horizontal muy desplazada arriba
    'D10': {'is_target': False, 'sq_left': False, 'sq_right': True,
            'sq_top': False, 'sq_bottom': False, 'h_offset': -12, 'v_asym': 5},
    # D11: cuadrado izquierdo + cuadrado abajo (diagonal confusa)
    'D11': {'is_target': False, 'sq_left': True,  'sq_right': False,
            'sq_top': False, 'sq_bottom': True,  'h_offset': 0,   'v_asym': -4},
    # D12: cruz desnuda + brazo vertical asimétrico largo arriba
    'D12': {'is_target': False, 'sq_left': False, 'sq_right': False,
            'sq_top': False, 'sq_bottom': False, 'h_offset': 0,   'v_asym': 14},
}

# Todos los distractores distribuidos con pesos equilibrados
# (ninguno domina; los tipo C [D7-D12] son un poco menos frecuentes por ser más difíciles)
DISTRACTOR_KEYS    = ['D1','D2','D3','D4','D5','D6','D7','D8','D9','D10','D11','D12']
DISTRACTOR_WEIGHTS = [ 10,  10,   9,   9,   9,   8,   9,   9,   8,    8,    6,    5]

# Dimensiones del canvas de estímulo
SW, SH = 68, 90    # ancho x alto del canvas
SQ = 6             # tamaño del cuadrado (px)
ARM_H = 20         # longitud brazo horizontal
ARM_V_UP   = 26    # longitud brazo vertical hacia arriba (base)
ARM_V_DOWN = 14    # longitud brazo vertical hacia abajo


def draw_plc_stimulus(canvas, stype, color_fg='#1A1A2E', color_bg='white',
                      w=SW, h=SH, line_width=2):
    """
    Dibuja el estímulo PLC en el canvas según su tipo.
    stype: dict con claves sq_left, sq_right, sq_top, sq_bottom,
                           h_offset (desplazamiento vertical de la línea H),
                           v_asym  (asimetría del brazo vertical superior)
    """
    canvas.delete('stim')
    canvas.config(bg=color_bg)

    cx = w // 2
    # Centro vertical base; h_offset desplaza la línea horizontal
    cy = h // 2 + stype.get('h_offset', 0)
    va = stype.get('v_asym', 0)   # extra px hacia arriba (+) o abajo (-)

    lw = line_width

    # ── Línea vertical ────────────────────────────────────────────────────
    top_y    = cy - ARM_V_UP - va
    bottom_y = cy + ARM_V_DOWN
    canvas.create_line(cx, top_y, cx, bottom_y,
                       width=lw, fill=color_fg, tags='stim', capstyle='round')

    # ── Línea horizontal ──────────────────────────────────────────────────
    left_x  = cx - ARM_H
    right_x = cx + ARM_H
    canvas.create_line(left_x, cy, right_x, cy,
                       width=lw, fill=color_fg, tags='stim', capstyle='round')

    # ── Cuadrados ─────────────────────────────────────────────────────────
    s = SQ
    def sq(x, y):
        canvas.create_rectangle(x - s//2, y - s//2, x + s//2, y + s//2,
                                 fill=color_fg, outline=color_fg, tags='stim')

    if stype.get('sq_left'):    sq(left_x,  cy)
    if stype.get('sq_right'):   sq(right_x, cy)
    if stype.get('sq_top'):     sq(cx, top_y)
    if stype.get('sq_bottom'):  sq(cx, bottom_y)


# =============================================================================
# PREDICTOR DE PERFIL COGNITIVO (MLP)
# =============================================================================

class PLC_MLPredictor:
    EXPECTED_FEATURES = [
        'age', 'TN', 'TA', 'O', 'C', 'TOT', 'CON', 'CP_percent',
        'total_time_seconds', 'mean_time_per_line', 'cv_time_percent',
        'processing_speed_cpm', 'efficiency_TA_per_min', 'FA', 'GQ',
        'fatigue_hits_percent', 'consistency_range',
        'block_1_hits', 'block_2_hits', 'block_3_hits', 'block_4_hits', 'block_5_hits',
        'omission_rate', 'commission_rate', 'accuracy_rate',
        'error_pattern_code', 'adjusted_score_by_age',
        'education_Primaria', 'education_Secundaria', 'education_Universitario',
        'hand_dominance_Derecha', 'hand_dominance_Izquierda'
    ]

    # Perfiles → descripción técnica sin patologías
    PROFILES = {
        0: {'key': 'Rendimiento_Tipico',    'nombre': 'Rendimiento Típico',
            'desc': 'Patrón de respuesta dentro del rango normativo esperado. Velocidad y precisión equilibradas.',
            'rasgos': ['CP% dentro del rango normativo', 'Proporción omisiones/comisiones equilibrada',
                       'Estabilidad temporal aceptable entre líneas']},
        1: {'key': 'Velocidad_Reducida',    'nombre': 'Velocidad Reducida con Omisiones',
            'desc': 'Patrón de procesamiento lento con umbral de respuesta elevado. Predominan omisiones sobre comisiones.',
            'rasgos': ['Alta tasa de omisiones', 'Velocidad por debajo de la media', 'Baja tasa de comisiones']},
        2: {'key': 'Alta_Velocidad',        'nombre': 'Alta Velocidad con Comisiones',
            'desc': 'Priorización de la velocidad sobre la exactitud. Umbral de respuesta bajo; tendencia a responder sin verificar.',
            'rasgos': ['Alta tasa de comisiones (falsos positivos)', 'Velocidad superior a la media',
                       'Trade-off velocidad-exactitud desfavorable']},
        3: {'key': 'Patron_Mixto',          'nombre': 'Patrón Mixto',
            'desc': 'Combinación de errores por omisión y comisión con alta variabilidad temporal. Inestabilidad en el control atencional.',
            'rasgos': ['Omisiones y comisiones elevadas simultáneamente', 'Alta variabilidad entre líneas',
                       'Posible decaimiento en el segundo bloque']},
        4: {'key': 'Decaimiento_Progresivo','nombre': 'Decaimiento Progresivo',
            'desc': 'Rendimiento inicial aceptable con deterioro progresivo. Baja resistencia a la monotonía o fatiga cognitiva.',
            'rasgos': ['Rendimiento menor en bloque 2 vs bloque 1', 'Aumento de errores hacia el final',
                       'Tiempo por línea creciente en últimas filas']},
        5: {'key': 'Fatiga_Alta',           'nombre': 'Activación Sostenida Baja',
            'desc': 'Rendimiento reducido y constante en todas las líneas. Posible estado de baja activación o fatiga previa.',
            'rasgos': ['Velocidad global reducida', 'Pocas oscilaciones entre bloques',
                       'CP% por debajo del percentil esperado']},
        6: {'key': 'Alto_Rendimiento',      'nombre': 'Alto Rendimiento',
            'desc': 'Patrón superior al normativo. Alta velocidad sostenida con precisión elevada.',
            'rasgos': ['CP% superior al percentil 85', 'Baja variabilidad temporal', 'Escasa fatiga entre bloques']},
        7: {'key': 'Inhibicion_Elevada',    'nombre': 'Estrategia Inhibitoria Elevada',
            'desc': 'Velocidad muy reducida con casi cero comisiones. Estrategia ultra-cautelosa o dificultad para liberar respuestas.',
            'rasgos': ['Muy baja tasa de comisiones', 'Velocidad por debajo del percentil 25',
                       'Alto tiempo por línea sin mejora entre bloques']},
    }

    def __init__(self):
        self.model  = None
        self.scaler = None
        self.model_info = {}
        self._load()

    def _load(self):
        if not TF_AVAILABLE:
            return
        mp = resource_path('d2_mlp_model_v3.keras')
        sp = resource_path('d2_scaler_v3.joblib')
        try:
            if os.path.exists(mp):
                self.model = load_model(mp)
                self.model_info = {'parameters': self.model.count_params(),
                                   'input_shape': self.model.input_shape,
                                   'output_shape': self.model.output_shape}
        except Exception as e:
            print(f'⚠️ Modelo: {e}')
        try:
            if os.path.exists(sp):
                self.scaler = joblib.load(sp)
        except Exception as e:
            print(f'⚠️ Scaler: {e}')

    def predict(self, data):
        if self.model is None:
            return {'model_used': False, 'error': 'Modelo no disponible'}
        try:
            age = data.get('age', 25)
            TN  = data.get('TN', 100);  TA = data.get('TA', 80)
            O   = data.get('O', 0);     C  = data.get('C', 0)
            TOT = O + C;                CON = TA - TOT
            CP  = (CON / TN * 100) if TN > 0 else 0
            tt  = data.get('total_time', 280)
            mtl = tt / 14
            cv  = data.get('cv_time', 10)
            ps  = (47*14/tt)*60 if tt > 0 else 0
            eff = (TA/tt)*60 if tt > 0 else 0
            FA  = TA/tt if tt > 0 else 0
            GQ  = (TA**2)/(TN*tt) if TN>0 and tt>0 else 0
            fh  = data.get('fatigue_hits', 0)
            cr  = data.get('consistency', 10)
            bh  = data.get('block_hits', [TA//5]*5)
            if len(bh) < 5: bh = [TA//5]*5
            or_ = (O/TN*100) if TN>0 else 0
            comr= (C/TN*100) if TN>0 else 0
            acc = (TA/TN*100) if TN>0 else 0
            ep  = 1 if O>C*2.5 else 2 if C>O*2.5 else 3 if O>C else 4 if C>O else 5
            af  = 0.92 if age<18 else 0.85 if age>60 else 0.95 if age>45 else 1.0
            edu = data.get('education','Universitario')
            hand= data.get('hand','Derecha')
            feat= [age,TN,TA,O,C,TOT,CON,CP,tt,mtl,cv,ps,eff,FA,GQ,fh,cr,
                   bh[0],bh[1],bh[2],bh[3],bh[4],or_,comr,acc,ep,CP*af,
                   1 if edu=='Primaria' else 0,
                   1 if edu=='Secundaria' else 0,
                   1 if edu=='Universitario' else 0,
                   1 if hand=='Derecha' else 0,
                   1 if hand=='Izquierda' else 0]
            X = np.array([feat])
            if self.scaler: X = self.scaler.transform(X)
            probs = self.model.predict(X, verbose=0)[0]
            pc    = int(np.argmax(probs))
            conf  = float(probs[pc])
            info  = self.PROFILES.get(pc, self.PROFILES[0])
            return {'model_used': True, 'predicted_code': pc,
                    'predicted_profile': info['key'],
                    'confidence': conf,
                    'confidence_percent': f'{conf*100:.1f}%',
                    'profile_info': info,
                    'all_probs': {self.PROFILES[i]['key']: float(p)
                                  for i, p in enumerate(probs)}}
        except Exception as e:
            return {'model_used': False, 'error': str(e)}


# =============================================================================
# APLICACIÓN PRINCIPAL
# =============================================================================

class PLCApp:

    C = {   # paleta
        'bg':        '#F4F6F9',
        'primary':   '#1A237E',
        'secondary': '#283593',
        'accent':    '#3949AB',
        'a_light':   '#E8EAF6',
        'white':     '#FFFFFF',
        'success':   '#2E7D32',
        's_bg':      '#E8F5E9',
        'warn':      '#E65100',
        'w_bg':      '#FFF3E0',
        'err':       '#B71C1C',
        'e_bg':      '#FFEBEE',
        'text':      '#1A1A2E',
        't_lite':    '#546E7A',
        'border':    '#C5CAE9',
        'sel':       '#3949AB',
        'sel_bg':    '#C5CAE9',
        'hdr':       '#FFFFFF',
        'chart1':    '#1565C0',
        'chart2':    '#E53935',
        'chart3':    '#F57F17',
        'chart4':    '#2E7D32',
    }

    def __init__(self, root):
        self.root = root
        self.root.title("PLC Professional — Evaluación Cognitiva")
        self.root.geometry("1440x900")
        self.root.configure(bg=self.C['bg'])
        self.root.resizable(True, True)
        try:
            from ctypes import windll
            windll.shcore.SetProcessDpiAwareness(1)
        except Exception:
            pass

        self.ml = PLC_MLPredictor()
        self.ml_pred = None
        self.results_dir = self._results_dir()

        # config test
        self.total_lines   = 14
        self.time_per_line = 20
        self.chars_per_line= 47

        # datos participante
        self.pid = self.pname = self.age = self.gender = ''
        self.education = self.hand = self.occupation = ''

        self.show_menu()

    # ─── UTILS ───────────────────────────────────────────────────────────────

    def _results_dir(self):
        base = os.path.join(os.path.expanduser('~'), 'Documents') \
               if sys.platform == 'win32' else os.path.expanduser('~')
        d = os.path.join(base, 'PLC_Resultados')
        os.makedirs(d, exist_ok=True)
        return d

    def clear(self):
        for w in self.root.winfo_children():
            w.destroy()

    def _scroll_frame(self):
        c = tk.Canvas(self.root, bg=self.C['bg'], highlightthickness=0)
        sb = ttk.Scrollbar(self.root, orient='vertical', command=c.yview)
        f  = tk.Frame(c, bg=self.C['bg'])
        f.bind('<Configure>', lambda e: c.configure(scrollregion=c.bbox('all')))
        c.create_window((0,0), window=f, anchor='nw')
        c.configure(yscrollcommand=sb.set)
        c.pack(side='left', fill='both', expand=True)
        sb.pack(side='right', fill='y')
        c.bind_all('<MouseWheel>',
                   lambda e: c.yview_scroll(int(-1*(e.delta/120)), 'units'))
        return f, c

    def _hdr(self, title, sub=None):
        b = tk.Frame(self.root, bg=self.C['primary'], pady=16)
        b.pack(fill='x')
        tk.Label(b, text=title, font=('Georgia', 20, 'bold'),
                 bg=self.C['primary'], fg=self.C['hdr']).pack()
        if sub:
            tk.Label(b, text=sub, font=('Helvetica', 10),
                     bg=self.C['primary'], fg='#C5CAE9').pack(pady=(2,0))

    def _btn(self, parent, text, cmd, style='primary',
             px=28, py=11, fs=12, side=None):
        palette = {
            'primary':   (self.C['accent'],   self.C['white']),
            'secondary': (self.C['secondary'],self.C['white']),
            'ghost':     (self.C['a_light'],  self.C['primary']),
            'danger':    (self.C['err'],      self.C['white']),
            'success':   (self.C['success'],  self.C['white']),
        }
        bg, fg = palette.get(style, palette['primary'])
        b = tk.Button(parent, text=text, command=cmd,
                      font=('Helvetica', fs, 'bold'),
                      bg=bg, fg=fg, padx=px, pady=py,
                      relief='flat', cursor='hand2')
        if side: b.pack(side=side, padx=6, pady=4)
        else:    b.pack(padx=6, pady=4)
        return b

    # =========================================================================
    # MENÚ PRINCIPAL
    # =========================================================================

    def show_menu(self):
        self.clear()
        self._hdr("PLC Professional", "Prueba de Líneas Cruzadas — Evaluación Cognitiva")
        C = self.C

        center = tk.Frame(self.root, bg=C['bg'])
        center.pack(expand=True, fill='both', padx=60, pady=30)

        # estado modelo
        ok = self.ml.model is not None
        sc = tk.Frame(center, bg=C['s_bg'] if ok else C['w_bg'],
                      padx=20, pady=12)
        sc.pack(fill='x', pady=(0,20))
        tk.Label(sc,
                 text=('● Módulo de perfil cognitivo activo' if ok
                       else '○ Modelo IA no disponible — solo métricas objetivas'),
                 font=('Helvetica', 11, 'bold'),
                 bg=C['s_bg'] if ok else C['w_bg'],
                 fg=C['success'] if ok else C['warn']).pack(side='left')

        # info
        ic = tk.Frame(center, bg=C['white'], padx=40, pady=28)
        ic.pack(fill='x', pady=(0,25))
        tk.Label(ic, text='Acerca de la prueba',
                 font=('Georgia', 15, 'bold'),
                 bg=C['white'], fg=C['primary']).pack(anchor='w', pady=(0,8))
        tk.Label(ic,
                 text=('Prueba de atención selectiva basada en identificación de estímulos objetivo\n'
                       'en un campo de distractores similares.  14 líneas · 20 s por línea · 47 estímulos/línea\n\n'
                       'Incluye fase de pre-prueba para verificar comprensión de instrucciones.'),
                 font=('Helvetica', 12), bg=C['white'], fg=C['text'],
                 justify='left', wraplength=860).pack(anchor='w')

        tk.Frame(center, bg=C['border'], height=1).pack(fill='x', pady=5)

        br = tk.Frame(center, bg=C['bg'])
        br.pack(pady=18)
        self._btn(br, '▶  Nueva Evaluación',   self.show_form,
                  style='primary', px=40, py=14, fs=14, side='left')
        self._btn(br, '📋  Historial',          self.show_history,
                  style='secondary', px=28, py=14, fs=13, side='left')

        tk.Label(center,
                 text='PLC Professional  v3.0  ·  Uso exclusivo para profesionales',
                 font=('Helvetica', 9), bg=C['bg'], fg=C['t_lite']
                 ).pack(side='bottom', pady=8)

    # =========================================================================
    # FORMULARIO SOCIODEMOGRÁFICO
    # =========================================================================

    def show_form(self):
        self.clear()
        self._hdr('Datos del Evaluado', 'Complete la información antes de iniciar')
        C = self.C

        inner, cv = self._scroll_frame()
        wrap = tk.Frame(inner, bg=C['bg'], padx=80, pady=25)
        wrap.pack(fill='both', expand=True)

        card = tk.Frame(wrap, bg=C['white'], padx=50, pady=35)
        card.pack(fill='x')

        tk.Label(card, text='Información del Participante',
                 font=('Georgia', 16, 'bold'),
                 bg=C['white'], fg=C['primary']).pack(anchor='w', pady=(0,18))

        def field(lbl, default=''):
            r = tk.Frame(card, bg=C['white']); r.pack(fill='x', pady=5)
            tk.Label(r, text=lbl, font=('Helvetica', 12, 'bold'),
                     bg=C['white'], fg=C['text'],
                     width=24, anchor='w').pack(side='left')
            e = tk.Entry(r, font=('Helvetica', 12),
                         bg=C['a_light'], fg=C['text'],
                         insertbackground=C['primary'], relief='flat')
            e.pack(side='left', fill='x', expand=True, ipady=7, padx=(8,0))
            e.insert(0, default)
            tk.Frame(r, bg=C['border'], height=1).pack(side='bottom', fill='x')
            return e

        def radios(lbl, opts, var):
            r = tk.Frame(card, bg=C['white']); r.pack(fill='x', pady=5)
            tk.Label(r, text=lbl, font=('Helvetica', 12, 'bold'),
                     bg=C['white'], fg=C['text'],
                     width=24, anchor='w').pack(side='left')
            f = tk.Frame(r, bg=C['white']); f.pack(side='left')
            for o in opts:
                tk.Radiobutton(f, text=o, variable=var, value=o,
                               bg=C['white'], fg=C['text'],
                               selectcolor=C['accent'],
                               activebackground=C['white'],
                               font=('Helvetica', 11)
                               ).pack(side='left', padx=10)

        self._id_e   = field('ID del Participante *')
        self._name_e = field('Nombre Completo *')
        self._age_e  = field('Edad *', '25')
        self._occ_e  = field('Ocupación / Cargo')

        tk.Frame(card, bg=C['border'], height=1).pack(fill='x', pady=10)

        self._gen_v  = tk.StringVar(value='Masculino')
        self._edu_v  = tk.StringVar(value='Universitario')
        self._hand_v = tk.StringVar(value='Derecha')

        radios('Género', ['Masculino','Femenino','Otro','No especificado'], self._gen_v)
        radios('Nivel Educativo', ['Primaria','Secundaria','Universitario','Posgrado'], self._edu_v)
        radios('Lateralidad', ['Derecha','Izquierda','Ambidiestro'], self._hand_v)

        br = tk.Frame(wrap, bg=C['bg']); br.pack(fill='x', pady=18)
        self._btn(br, '← Menú', self.show_menu, style='ghost', side='left')
        self._btn(br, 'Continuar a pre-prueba  →', self._validate_form,
                  style='primary', px=36, py=13, fs=13, side='right')

        inner.update_idletasks(); cv.configure(scrollregion=cv.bbox('all'))

    def _validate_form(self):
        pid   = self._id_e.get().strip()
        pname = self._name_e.get().strip()
        age_s = self._age_e.get().strip()
        if not pid or not pname:
            messagebox.showerror('Campos requeridos',
                                 'ID y Nombre son obligatorios.')
            return
        if not age_s.isdigit() or not (5 <= int(age_s) <= 100):
            messagebox.showerror('Edad inválida',
                                 'Ingrese una edad entre 5 y 100 años.')
            return
        self.pid        = pid
        self.pname      = pname
        self.age        = age_s
        self.occupation = self._occ_e.get().strip()
        self.gender     = self._gen_v.get()
        self.education  = self._edu_v.get()
        self.hand       = self._hand_v.get()
        self.show_preprueba()

    # =========================================================================
    # PRE-PRUEBA — igual al ejemplo de la imagen
    # Muestra: target etiquetado + 3 distractores etiquetados
    # El evaluado debe marcar correctamente antes de empezar
    # =========================================================================

    def show_preprueba(self):
        self.clear()
        self._hdr('Pre-Prueba — Ejemplo',
                  'Identifique el estímulo objetivo antes de comenzar la prueba real')
        C = self.C

        inner, cv = self._scroll_frame()
        wrap = tk.Frame(inner, bg=C['bg'], padx=40, pady=20)
        wrap.pack(fill='both', expand=True)

        # ── Título ────────────────────────────────────────────────────────
        tk.Label(wrap, text='EJEMPLO: LÍNEAS CRUZADAS',
                 font=('Georgia', 17, 'bold'),
                 bg=C['bg'], fg=C['primary']).pack(anchor='w', pady=(0,12))

        demo_row = tk.Frame(wrap, bg=C['white'], padx=30, pady=24)
        demo_row.pack(fill='x', pady=(0,10))

        # ─ TARGET ─────────────────────────────────────────────────────────
        left = tk.Frame(demo_row, bg=C['white'])
        left.pack(side='left', padx=30)
        tk.Label(left, text='ESTÍMULO OBJETIVO\n(TARGET)',
                 font=('Helvetica', 11, 'bold'),
                 bg=C['white'], fg=C['primary'], justify='center').pack()
        tk.Label(left, text='→', font=('Helvetica', 24),
                 bg=C['white'], fg=C['primary']).pack(side='left', padx=5)

        t_c = tk.Canvas(left, width=SW+20, height=SH+20,
                        bg=C['s_bg'], highlightthickness=2,
                        highlightbackground=C['success'])
        t_c.pack(side='left', padx=5)
        draw_plc_stimulus(t_c, STIM_TYPES['T'],
                          color_fg=C['text'], color_bg=C['s_bg'],
                          w=SW+20, h=SH+20, line_width=3)

        # separador vertical
        tk.Frame(demo_row, bg=C['border'], width=1).pack(
            side='left', fill='y', padx=20)

        # ─ DISTRACTORES ───────────────────────────────────────────────────
        right = tk.Frame(demo_row, bg=C['white'])
        right.pack(side='left', padx=10)
        tk.Label(right, text='DISTRACTORES (NO MARCAR)',
                 font=('Helvetica', 11, 'bold'),
                 bg=C['white'], fg=C['err']).pack(anchor='w', pady=(0,8))

        dist_row = tk.Frame(right, bg=C['white'])
        dist_row.pack()
        for i, dk in enumerate(['D1','D2','D3'], start=1):
            df = tk.Frame(dist_row, bg=C['white'])
            df.pack(side='left', padx=18)
            tk.Label(df, text=f'{i})',
                     font=('Helvetica', 12, 'bold'),
                     bg=C['white'], fg=C['t_lite']).pack()
            dc = tk.Canvas(df, width=SW+20, height=SH+20,
                           bg=C['e_bg'], highlightthickness=1,
                           highlightbackground=C['err'])
            dc.pack()
            draw_plc_stimulus(dc, STIM_TYPES[dk],
                              color_fg=C['text'], color_bg=C['e_bg'],
                              w=SW+20, h=SH+20, line_width=2)

        # ── Instrucción ───────────────────────────────────────────────────
        instr = tk.Frame(wrap, bg=C['a_light'], padx=24, pady=16)
        instr.pack(fill='x', pady=10)
        tk.Label(instr,
                 text='INSTRUCCIÓN AL EVALUADO',
                 font=('Helvetica', 12, 'bold'),
                 bg=C['a_light'], fg=C['primary']).pack(anchor='w')
        tk.Label(instr,
                 text=('Marque todos los estímulos que sean IGUALES al objetivo (TARGET): '
                       'una cruz con un cuadrado negro a CADA LADO del brazo horizontal.\n'
                       'No marque los distractores aunque tengan forma similar.'),
                 font=('Helvetica', 11),
                 bg=C['a_light'], fg=C['text'],
                 wraplength=900, justify='left').pack(anchor='w', pady=(5,0))

        # ── Mini-práctica: 8 estímulos, el evaluado debe marcar los correctos ──
        tk.Label(wrap,
                 text='Práctica: marque los estímulos correctos en la fila de abajo',
                 font=('Helvetica', 12, 'bold'),
                 bg=C['bg'], fg=C['secondary']).pack(anchor='w', pady=(14,4))

        NEED = 2
        self._pre_need = NEED

        # ── Contenedor de la práctica (se reconstruirá al repetir) ────────
        self._pre_container = tk.Frame(wrap, bg=C['bg'])
        self._pre_container.pack(fill='x')

        self._pre_fb = tk.Label(wrap, text='',
                                font=('Helvetica', 12, 'bold'),
                                bg=C['bg'], fg=C['success'])
        self._pre_fb.pack(pady=4)

        prog_row = tk.Frame(wrap, bg=C['bg']); prog_row.pack(pady=4)
        tk.Label(prog_row, text='Marcados correctos:',
                 font=('Helvetica', 11),
                 bg=C['bg'], fg=C['t_lite']).pack(side='left')
        self._pre_pvar = tk.DoubleVar(value=0)
        ttk.Progressbar(prog_row, variable=self._pre_pvar,
                        maximum=NEED, length=200).pack(side='left', padx=8)
        self._pre_cnt_lbl = tk.Label(prog_row,
                                     text=f'0 / {NEED}',
                                     font=('Helvetica', 12, 'bold'),
                                     bg=C['bg'], fg=C['primary'])
        self._pre_cnt_lbl.pack(side='left')

        br = tk.Frame(wrap, bg=C['bg']); br.pack(fill='x', pady=14)
        self._btn(br, '← Volver', self.show_form, style='ghost', side='left')

        # Botón «Repetir práctica» — centrado, siempre activo
        self._retry_btn = tk.Button(br,
                                    text='🔄  Repetir práctica',
                                    font=('Helvetica', 11, 'bold'),
                                    bg=C['warn'], fg=C['white'],
                                    padx=20, pady=10, relief='flat',
                                    cursor='hand2',
                                    command=self._reset_practice)
        self._retry_btn.pack(side='left', padx=10)

        self._start_btn = tk.Button(br,
                                    text='Iniciar Prueba  →',
                                    font=('Helvetica', 13, 'bold'),
                                    bg=C['border'], fg=C['t_lite'],
                                    padx=36, pady=12, relief='flat',
                                    cursor='arrow', state='disabled',
                                    command=self.start_test)
        self._start_btn.pack(side='right', padx=6)

        # Render inicial de la práctica
        self._reset_practice()

        inner.update_idletasks(); cv.configure(scrollregion=cv.bbox('all'))

    def _reset_practice(self):
        """Genera una nueva tanda aleatoria de estímulos de práctica y re-dibuja el contenedor."""
        # Limpiar contenedor anterior
        for w in self._pre_container.winfo_children():
            w.destroy()

        self._pre_ok   = 0
        self._pre_btns = []
        self._pre_pvar.set(0)
        self._pre_cnt_lbl.config(text=f'0 / {self._pre_need}')
        self._pre_fb.config(text='', fg=self.C['success'])
        self._start_btn.config(state='disabled',
                               bg=self.C['border'], fg=self.C['t_lite'], cursor='arrow')

        # Nueva fila aleatoria garantizando exactamente 2 targets
        prow = tk.Frame(self._pre_container, bg=self.C['bg'])
        prow.pack(pady=5)

        self._pre_line = self._gen_practice_line(10)
        for idx, sinfo in enumerate(self._pre_line):
            self._make_pre_stim(prow, idx, sinfo)

    def _gen_practice_line(self, n=10):
        """Genera una línea de práctica con exactamente 2 targets y el resto distractores variados."""
        # Siempre exactamente 2 targets
        targets = ['T', 'T']
        # Elegir (n-2) distractores asegurando que usen al menos 4 tipos distintos
        all_dist = random.choices(DISTRACTOR_KEYS, DISTRACTOR_WEIGHTS, k=n-2)
        # Garantizar variedad: si hay menos de 4 tipos distintos, reemplazar aleatoriamente
        while len(set(all_dist)) < min(4, n-2):
            i = random.randrange(len(all_dist))
            all_dist[i] = random.choice(DISTRACTOR_KEYS)
        combined = targets + all_dist
        random.shuffle(combined)
        return [{'key': k, **STIM_TYPES[k]} for k in combined]

    def _make_pre_stim(self, parent, idx, sinfo):
        C = self.C
        f = tk.Frame(parent, bg=C['bg'], padx=2); f.pack(side='left')
        cv = tk.Canvas(f, width=SW+10, height=SH+10,
                       bg=C['white'], highlightthickness=1,
                       highlightbackground=C['border'])
        cv.pack()
        draw_plc_stimulus(cv, sinfo,
                          color_fg=C['text'], color_bg=C['white'],
                          w=SW+10, h=SH+10)
        cv.bind('<Button-1>',
                lambda e, i=idx, c=cv, s=sinfo: self._pre_toggle(i, c, s))
        self._pre_btns.append({'cv': cv, 'info': sinfo, 'sel': False})

    def _pre_toggle(self, idx, cv, sinfo):
        C   = self.C
        btn = self._pre_btns[idx]
        is_t = sinfo['is_target']
        if btn['sel']:
            btn['sel'] = False
            cv.config(bg=C['white'], highlightbackground=C['border'])
            draw_plc_stimulus(cv, sinfo, C['text'], C['white'], SW+10, SH+10)
            if is_t: self._pre_ok = max(0, self._pre_ok - 1)
        else:
            btn['sel'] = True
            if is_t:
                cv.config(bg=C['s_bg'], highlightbackground=C['success'])
                draw_plc_stimulus(cv, sinfo, C['success'], C['s_bg'], SW+10, SH+10)
                self._pre_ok += 1
                self._pre_fb.config(
                    text='✓ Correcto — ese es el estímulo objetivo.',
                    fg=C['success'])
            else:
                cv.config(bg=C['e_bg'], highlightbackground=C['err'])
                draw_plc_stimulus(cv, sinfo, C['err'], C['e_bg'], SW+10, SH+10)
                self._pre_fb.config(
                    text='✗ Ese no es el objetivo. Observe bien el cuadrado en AMBOS lados.',
                    fg=C['err'])

        prog = min(self._pre_ok, self._pre_need)
        self._pre_pvar.set(prog)
        self._pre_cnt_lbl.config(text=f'{self._pre_ok} / {self._pre_need}')

        if self._pre_ok >= self._pre_need:
            self._start_btn.config(state='normal',
                                   bg=C['accent'], fg=C['white'], cursor='hand2')
            self._pre_fb.config(
                text='✓ ¡Perfecto! Ya puede iniciar la prueba.',
                fg=C['success'])
        else:
            self._start_btn.config(state='disabled',
                                   bg=C['border'], fg=C['t_lite'], cursor='arrow')

    # =========================================================================
    # GENERACIÓN DE LÍNEAS DEL TEST
    # =========================================================================

    def start_test(self):
        self.current_line = 0
        self.lines_data   = []
        self.test_lines   = []
        self.line_times   = []
        self.click_log    = []   # timestamp ms por ítem

        for li in range(self.total_lines):
            n_t = random.randint(8, 12)
            positions = list(range(self.chars_per_line))
            random.shuffle(positions)
            tpos = set(positions[:n_t])
            line = []
            for pos in range(self.chars_per_line):
                if pos in tpos:
                    s = {'key':'T', **STIM_TYPES['T'], 'pos': pos}
                else:
                    k = random.choices(DISTRACTOR_KEYS, DISTRACTOR_WEIGHTS)[0]
                    s = {'key': k, **STIM_TYPES[k], 'pos': pos}
                line.append(s)
            self.test_lines.append(line)

        self.show_line()

    # =========================================================================
    # PANTALLA DE LÍNEA — 2 FILAS máximo (≈24 + 23 estímulos)
    # =========================================================================

    def show_line(self):
        self.clear()
        if self.current_line >= self.total_lines:
            self.finish_test(); return

        C   = self.C
        ld  = self.test_lines[self.current_line]
        N   = self.chars_per_line   # 47

        # ── ENCABEZADO ────────────────────────────────────────────────────
        hdr = tk.Frame(self.root, bg=C['primary'], pady=10)
        hdr.pack(fill='x')
        tk.Label(hdr,
                 text=f'LÍNEA  {self.current_line+1} / {self.total_lines}',
                 font=('Helvetica', 17, 'bold'),
                 bg=C['primary'], fg=C['hdr']).pack(side='left', padx=20)

        self._status_lbl = tk.Label(hdr, text='Marcados: 0',
                                    font=('Helvetica', 13),
                                    bg=C['primary'], fg='#C5CAE9')
        self._status_lbl.pack(side='left', padx=25)

        self._timer_lbl = tk.Label(hdr, text='20.0 s',
                                   font=('Helvetica', 21, 'bold'),
                                   bg=C['primary'], fg=C['hdr'])
        self._timer_lbl.pack(side='right', padx=22)

        # ── BARRA TEMPORAL ────────────────────────────────────────────────
        bframe = tk.Frame(self.root, bg=C['bg'], pady=3)
        bframe.pack(fill='x', padx=16)

        sty = ttk.Style(); sty.theme_use('clam')
        for nm, bg_col in [('TBlue.Horizontal.TProgressbar',  C['accent']),
                            ('TWarn.Horizontal.TProgressbar',  C['warn']),
                            ('TCrit.Horizontal.TProgressbar',  C['err'])]:
            sty.configure(nm, troughcolor=C['a_light'],
                          background=bg_col, thickness=12)

        self._pvar = tk.DoubleVar(value=100)
        self._pbar = ttk.Progressbar(bframe, variable=self._pvar,
                                     maximum=100, style='TBlue.Horizontal.TProgressbar')
        self._pbar.pack(fill='x')

        # ── RECORDATORIO BREVE ────────────────────────────────────────────
        tk.Label(self.root,
                 text='  Marque las cruces con cuadrado negro en AMBOS extremos del brazo horizontal  ■—+—■',
                 font=('Helvetica', 10, 'bold'),
                 bg=C['secondary'], fg=C['hdr'], pady=5).pack(fill='x')

        # ── ÁREA DE ESTÍMULOS — 2 FILAS, spacing dinámico ────────────────
        stim_area = tk.Frame(self.root, bg=C['bg'])
        stim_area.pack(expand=True, fill='both', padx=6, pady=4)

        self._char_btns      = []
        self._current_sels   = set()

        # Calcular ancho disponible y dividir en 2 filas equilibradas
        screen_w = self.root.winfo_screenwidth()
        # Espacio disponible (descontando márgenes)
        available_w = screen_w - 40
        # Calcular padding horizontal dinámico entre estímulos
        n_per_row   = 24
        item_total_w = SW + 4  # canvas + padding
        total_used   = n_per_row * item_total_w
        dyn_padx = max(1, min(4, (available_w - total_used) // (n_per_row * 2)))

        # Dividir 47 en 2 filas: 24 + 23
        split = 24
        rows  = [ld[:split], ld[split:]]

        for row_idx, row_items in enumerate(rows):
            start_idx = 0 if row_idx == 0 else split
            rf = tk.Frame(stim_area, bg=C['bg'])
            rf.pack(expand=True, pady=5)
            for rel_idx, sinfo in enumerate(row_items):
                abs_idx = start_idx + rel_idx
                self._make_test_stim(rf, abs_idx, sinfo, dyn_padx)

        # ── BARRA INFERIOR: puntos de progreso + botón "Siguiente" ────────
        bot = tk.Frame(self.root, bg=C['bg'], pady=4)
        bot.pack(fill='x', padx=14)

        dots = tk.Frame(bot, bg=C['bg']); dots.pack(side='left')
        for i in range(self.total_lines):
            col = C['accent'] if i < self.current_line else \
                  C['success'] if i == self.current_line else C['border']
            tk.Label(dots, text='●', font=('Helvetica', 8),
                     bg=C['bg'], fg=col).pack(side='left', padx=1)

        # Botón siguiente — pequeño, esquina derecha (respaldo manual)
        tk.Button(bot, text='Siguiente →',
                  font=('Helvetica', 9),
                  bg=C['a_light'], fg=C['primary'],
                  padx=12, pady=5, relief='flat', cursor='hand2',
                  command=self.next_line).pack(side='right', padx=4)

        self._line_start = time.time()
        self._timer_running = True
        self._update_timer()

    def _make_test_stim(self, parent, idx, sinfo, padx=2):
        C = self.C
        # Hitbox = canvas completo (SW×SH = 68×90) — área de clic ya es >= 40×40
        f  = tk.Frame(parent, bg=C['bg'], padx=padx)
        f.pack(side='left')
        cv = tk.Canvas(f, width=SW, height=SH,
                       bg=C['white'], highlightthickness=1,
                       highlightbackground=C['border'],
                       cursor='hand2')
        cv.pack()
        draw_plc_stimulus(cv, sinfo, C['text'], C['white'])
        # El bind cubre TODO el canvas (hitbox = 68×90 px, suficiente margen motor)
        cv.bind('<Button-1>',
                lambda e, i=idx, c=cv, s=sinfo: self._toggle(i, c, s))
        self._char_btns.append({'cv': cv, 'info': sinfo, 'sel': False})

    def _toggle(self, idx, cv, sinfo):
        C   = self.C
        btn = self._char_btns[idx]
        now = time.time()
        # registrar timestamp
        self.click_log.append({
            'line':      self.current_line + 1,
            'stim_idx':  idx,
            'is_target': sinfo['is_target'],
            'stim_key':  sinfo['key'],
            'action':    'desel' if btn['sel'] else 'sel',
            'elapsed_ms': round((now - self._line_start) * 1000)
        })
        if btn['sel']:
            btn['sel'] = False
            self._current_sels.discard(idx)
            # Quitar selección: borde normal, fondo blanco, figura en negro
            cv.config(bg=C['white'], highlightthickness=1,
                      highlightbackground=C['border'])
            draw_plc_stimulus(cv, sinfo, C['text'], C['white'])
        else:
            btn['sel'] = True
            self._current_sels.add(idx)
            # Selección NO INVASIVA: solo engrosamos el borde azul claro
            # La figura queda NEGRA y visible — solo el marco cambia
            cv.config(bg=C['white'], highlightthickness=3,
                      highlightbackground='#1565C0')
            draw_plc_stimulus(cv, sinfo, C['text'], C['white'])
        self._status_lbl.config(text=f'Marcados: {len(self._current_sels)}')

    def _update_timer(self):
        if not self._timer_running: return
        elapsed   = time.time() - self._line_start
        remaining = max(0.0, self.time_per_line - elapsed)
        pct       = (remaining / self.time_per_line) * 100
        C = self.C

        self._timer_lbl.config(text=f'{remaining:.1f} s')
        self._pvar.set(pct)

        if remaining > 10:
            self._timer_lbl.config(fg=C['hdr'])
            self._pbar.config(style='TBlue.Horizontal.TProgressbar')
        elif remaining > 5:
            self._timer_lbl.config(fg='#FFD54F')
            self._pbar.config(style='TWarn.Horizontal.TProgressbar')
        else:
            self._timer_lbl.config(fg='#EF9A9A')
            self._pbar.config(style='TCrit.Horizontal.TProgressbar')

        if remaining > 0:
            self.root.after(100, self._update_timer)
        else:
            self._timer_running = False
            # Transición suave: flash blanco 500ms antes de cargar siguiente línea
            self.root.configure(bg='#FFFFFF')
            self.root.after(500, self.next_line)

    def next_line(self):
        self._timer_running = False
        lt = time.time() - self._line_start
        hits = oms = coms = targets = 0
        for b in self._char_btns:
            it = b['info']['is_target']
            sl = b['sel']
            if it:
                targets += 1
                if sl: hits += 1
                else:  oms  += 1
            elif sl:
                coms += 1

        self.lines_data.append({
            'linea': self.current_line+1,
            'targets_total': targets,
            'aciertos':   hits,
            'omisiones':  oms,
            'comisiones': coms,
            'tiempo_s':   round(lt, 3),
            'tiempo_pct': round((min(lt, self.time_per_line)/self.time_per_line)*100, 1)
        })
        self.line_times.append(lt)
        self.current_line += 1
        if self.current_line < self.total_lines:
            self.show_line()
        else:
            self.finish_test()

    # =========================================================================
    # CÁLCULO DE MÉTRICAS
    # =========================================================================

    def finish_test(self):
        self.ml_pred = None
        self._calc_metrics()
        self._run_ml()
        self._save_excel()
        self.show_results()

    def _calc_metrics(self):
        ld = self.lines_data
        self.TA  = sum(l['aciertos']   for l in ld)
        self.O   = sum(l['omisiones']  for l in ld)
        self.COM = sum(l['comisiones'] for l in ld)   # COM = Comisiones (C reservada para paleta)
        self.TN  = self.TA + self.O
        self.TOT = self.O + self.COM
        self.CON = self.TA - self.TOT
        self.CP  = (self.CON / self.TN * 100) if self.TN > 0 else 0

        times = [l['tiempo_s'] for l in ld]
        self.total_time     = sum(times)
        self.mean_tpl       = statistics.mean(times)
        self.std_tpl        = statistics.stdev(times) if len(times)>1 else 0
        self.cv_time        = (self.std_tpl/self.mean_tpl*100) if self.mean_tpl>0 else 0

        self.proc_speed     = (self.chars_per_line*self.total_lines/self.total_time)*60 \
                               if self.total_time>0 else 0
        self.efficiency     = (self.TA/self.total_time)*60 if self.total_time>0 else 0
        self.FA             = self.TA/self.total_time if self.total_time>0 else 0
        self.GQ             = (self.TA**2)/(self.TN*self.total_time) \
                               if self.TN>0 and self.total_time>0 else 0

        hits = [l['aciertos'] for l in ld]
        hm   = statistics.mean(hits)
        hs   = statistics.stdev(hits) if len(hits)>1 else 0
        self.VAR         = (hs/hm*100) if hm>0 else 0
        self.estabilidad = max(0, 100-self.VAR)
        self.consistency = max(hits)-min(hits)

        mid = len(ld)//2
        h1  = sum(l['aciertos'] for l in ld[:mid])
        h2  = sum(l['aciertos'] for l in ld[mid:])
        self.TRM = ((h2-h1)/h1*100) if h1>0 else 0

        self.IVR = (self.COM/(self.proc_speed/100)) if self.proc_speed>0 else 0

        # bloques (5) para MLP
        self.block_hits = []
        bs = len(ld)//5
        for i in range(5):
            s = i*bs; e = s+bs if i<4 else len(ld)
            self.block_hits.append(sum(l['aciertos'] for l in ld[s:e]))
        while len(self.block_hits)<5: self.block_hits.append(self.TA//5)

        self.error_pat = ('Predominio Omisiones'   if self.O   > self.COM*3 else
                          'Predominio Comisiones'  if self.COM > self.O*3   else
                          'Mayor Omisiones'        if self.O   > self.COM   else
                          'Mayor Comisiones'       if self.COM > self.O     else 'Mixto')

        age = int(self.age) if str(self.age).isdigit() else 25
        af  = 0.90 if age<18 else 0.85 if age>60 else 0.95 if age>40 else 1.0
        self.adj_score = self.CP * af

        # tiempo de reacción estimado por ítem (del click_log)
        first_clicks = {}
        for cl in self.click_log:
            if cl['action']=='sel':
                key = (cl['line'], cl['stim_idx'])
                if key not in first_clicks:
                    first_clicks[key] = cl['elapsed_ms']
        rts = list(first_clicks.values())
        self.mean_rt_ms = statistics.mean(rts) if rts else 0
        self.med_rt_ms  = statistics.median(rts) if rts else 0

        # estilo de atención dominante
        if self.O > self.COM * 2:
            self.attn_style = 'Cauteloso / Lento'
            self.attn_desc  = ('Predominan las omisiones sobre las comisiones. '
                               'El evaluado tiende a ser selectivo y lento, dejando pasar estímulos objetivo.')
        elif self.COM > self.O * 2:
            self.attn_style = 'Impulsivo / Rápido'
            self.attn_desc  = ('Predominan las comisiones. El evaluado responde rápido con bajo umbral de selección, '
                               'marcando estímulos distractores.')
        elif self.CP >= 75:
            self.attn_style = 'Eficiente / Equilibrado'
            self.attn_desc  = 'Buen balance entre velocidad y precisión. Pocos errores en ambas direcciones.'
        else:
            self.attn_style = 'Mixto / Inestable'
            self.attn_desc  = 'Patrón mixto de errores. Alta variabilidad en la ejecución entre líneas.'

    def _run_ml(self):
        self.ml_pred = self.ml.predict({
            'age': int(self.age) if str(self.age).isdigit() else 25,
            'education': self.education, 'hand': self.hand,
            'TN': self.TN, 'TA': self.TA, 'O': self.O, 'C': self.COM,
            'total_time': self.total_time, 'cv_time': self.cv_time,
            'fatigue_hits': self.TRM, 'consistency': self.consistency,
            'block_hits': self.block_hits
        })

    # =========================================================================
    # EXCEL PROFESIONAL — 5 HOJAS
    # =========================================================================

    def _save_excel(self):
        try:
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.chart import LineChart, BarChart, Reference
            ts   = datetime.now().strftime('%Y%m%d_%H%M%S')
            fn   = f'PLC_{self.pid}_{ts}.xlsx'
            fp   = os.path.join(self.results_dir, fn)

            hfill = PatternFill('solid', fgColor='1A237E')
            hfont = Font(color='FFFFFF', bold=True, name='Calibri', size=11)
            halign = Alignment(horizontal='center', vertical='center', wrap_text=True)

            def style_hdr(ws, row=1):
                for cell in ws[row]:
                    if cell.value is not None:
                        cell.fill = hfill; cell.font = hfont; cell.alignment = halign

            def set_w(ws, widths):
                from openpyxl.utils import get_column_letter
                for i,w in enumerate(widths,1):
                    ws.column_dimensions[get_column_letter(i)].width = w

            with pd.ExcelWriter(fp, engine='openpyxl') as writer:

                # ── H1: Resumen ───────────────────────────────────────────
                d1 = {
                    'Campo': ['─ IDENTIFICACIÓN ─','ID','Nombre','Edad','Género',
                              'Educación','Lateralidad','Ocupación','Fecha',
                              '','─ MÉTRICAS ESTÁNDAR ─',
                              'TN','TA','O','C','TOT','CON','CP %',
                              '','─ MÉTRICAS EXTENDIDAS ─',
                              'Velocidad (estím/min)','Eficiencia (aciert/min)',
                              'FA','VAR %','Estabilidad %','TRM %','IVR',
                              'Patrón de error','Score ajustado edad',
                              '','─ TIEMPO DE REACCIÓN ─',
                              'TR medio (ms)','TR mediana (ms)',
                              '','─ ESTILO ATENCIONAL ─',
                              'Estilo dominante','Descripción'],
                    'Valor': ['',self.pid,self.pname,self.age,self.gender,
                              self.education,self.hand,
                              getattr(self,'occupation',''),
                              datetime.now().strftime('%Y-%m-%d %H:%M'),
                              '','',
                              self.TN,self.TA,self.O,self.COM,self.TOT,self.CON,
                              round(self.CP,2),
                              '','',
                              round(self.proc_speed,1),round(self.efficiency,1),
                              round(self.FA,4),round(self.VAR,2),
                              round(self.estabilidad,1),round(self.TRM,2),
                              round(self.IVR,3),self.error_pat,round(self.adj_score,2),
                              '','',
                              round(self.mean_rt_ms,1),round(self.med_rt_ms,1),
                              '','',
                              self.attn_style,self.attn_desc]
                }
                pd.DataFrame(d1).to_excel(writer, sheet_name='01_Resumen', index=False)
                ws1 = writer.sheets['01_Resumen']
                style_hdr(ws1); set_w(ws1,[38,55])

                # ── H2: Análisis líneas ───────────────────────────────────
                df2 = pd.DataFrame(self.lines_data)
                df2['precision_%'] = (df2['aciertos']/df2['targets_total'].replace(0,1)*100).round(1)
                df2['tasa_om_%']   = (df2['omisiones']/df2['targets_total'].replace(0,1)*100).round(1)
                df2['tasa_com_%']  = (df2['comisiones']/df2['targets_total'].replace(0,1)*100).round(1)
                df2.to_excel(writer, sheet_name='02_Analisis_Lineas', index=False)
                ws2 = writer.sheets['02_Analisis_Lineas']
                style_hdr(ws2); set_w(ws2,[8,14,10,11,11,10,12,10,10,10])

                # ── H3: Micro-tendencia ───────────────────────────────────
                df3 = pd.DataFrame(self.click_log) if self.click_log else \
                      pd.DataFrame(columns=['line','stim_idx','is_target',
                                            'stim_key','action','elapsed_ms'])
                df3.to_excel(writer, sheet_name='03_Micro_Tendencia', index=False)
                ws3 = writer.sheets['03_Micro_Tendencia']
                style_hdr(ws3); set_w(ws3,[8,10,10,10,8,12])

                # ── H4: Perfil de rasgos ──────────────────────────────────
                if self.ml_pred and self.ml_pred.get('model_used'):
                    pi   = self.ml_pred['profile_info']
                    pn   = pi['nombre']
                    pdesc= pi['desc']
                    pras = '\n'.join(f'• {r}' for r in pi['rasgos'])
                    pconf= self.ml_pred['confidence_percent']
                    ptop = '  |  '.join(
                        f"{k.replace('_',' ')}: {v*100:.1f}%"
                        for k,v in sorted(self.ml_pred['all_probs'].items(),
                                          key=lambda x:x[1],reverse=True)[:3])
                else:
                    pn=pdesc=pras=pconf=ptop='No disponible (modelo no cargado)'

                narrative = self._narrative()
                d4 = {
                    'Dimensión':[
                        '─ NARRATIVA TÉCNICA ─','Descripción integrada','',
                        '─ ESTILO ATENCIONAL ─','Estilo dominante','Descripción','',
                        '─ PERFIL COGNITIVO (IA) ─','Perfil','Confianza','Descripción',
                        'Rasgos','Perfiles top-3','',
                        '─ ALCANCE ─','Nota para el evaluador'],
                    'Contenido':[
                        '',narrative,'',
                        '',self.attn_style,self.attn_desc,'',
                        '',pn,pconf,pdesc,pras,ptop,'',
                        '',('Este informe describe indicadores del rendimiento cognitivo durante '
                            'la prueba PLC. No constituye un diagnóstico clínico. '
                            'El evaluador debe integrar estos datos con historia clínica, '
                            'entrevista y otros instrumentos.')
                    ]
                }
                pd.DataFrame(d4).to_excel(writer, sheet_name='04_Perfil_Rasgos', index=False)
                ws4 = writer.sheets['04_Perfil_Rasgos']
                style_hdr(ws4); set_w(ws4,[30,90])
                for row in ws4.iter_rows(min_row=2,min_col=2,max_col=2):
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True,vertical='top')

                # ── H5: Gráficos ──────────────────────────────────────────
                gd = {
                    'Linea':      [l['linea']    for l in self.lines_data],
                    'Aciertos':   [l['aciertos'] for l in self.lines_data],
                    'Omisiones':  [l['omisiones']for l in self.lines_data],
                    'Comisiones': [l['comisiones']for l in self.lines_data],
                    'Tiempo_s':   [round(l['tiempo_s'],2) for l in self.lines_data],
                }
                pd.DataFrame(gd).to_excel(writer, sheet_name='05_Graficos', index=False)
                ws5 = writer.sheets['05_Graficos']
                style_hdr(ws5); set_w(ws5,[8,11,11,12,10])
                try:
                    # Curva de aciertos
                    lc = LineChart(); lc.title='Curva de Aciertos por Línea'
                    lc.style=10; lc.height=12; lc.width=22
                    dr = Reference(ws5,min_col=2,min_row=1,max_row=15)
                    lc.add_data(dr,titles_from_data=True)
                    lc.set_categories(Reference(ws5,min_col=1,min_row=2,max_row=15))
                    ws5.add_chart(lc,'G2')
                    # Barras errores
                    bc = BarChart(); bc.type='col'; bc.title='Errores por Línea'
                    bc.style=10; bc.height=12; bc.width=22
                    er = Reference(ws5,min_col=3,min_row=1,max_col=4,max_row=15)
                    bc.add_data(er,titles_from_data=True)
                    bc.set_categories(Reference(ws5,min_col=1,min_row=2,max_row=15))
                    ws5.add_chart(bc,'G22')
                except Exception:
                    pass

            self.excel_file = fp
            print(f'✅ Excel guardado: {fp}')
        except PermissionError:
            messagebox.showwarning(
                'Archivo en uso',
                f'El archivo Excel ya está abierto en otro programa.\n'
                f'Por favor ciérrelo e intente de nuevo.\n\n{fp}')
            self.excel_file = None
        except Exception as e:
            import traceback; traceback.print_exc()
            messagebox.showerror('Error Excel', str(e))
            self.excel_file = None

    def _narrative(self):
        parts = []
        parts.append(f'Velocidad de procesamiento: {self.proc_speed:.0f} estím/min '
                     f'({"superior" if self.proc_speed>120 else "dentro de" if self.proc_speed>80 else "por debajo de"} la media).')
        parts.append(f'Índice de concentración CP = {self.CP:.1f} % '
                     f'({"alto" if self.CP>=80 else "medio" if self.CP>=60 else "bajo"}).')
        parts.append(f'Estabilidad del rastro visual: {self.estabilidad:.0f} % (VAR = {self.VAR:.1f} %).')
        trm_txt = ('mejora' if self.TRM>5 else
                   'estable' if self.TRM>=-15 else
                   'decaimiento moderado' if self.TRM>=-30 else
                   'decaimiento significativo')
        parts.append(f'Resistencia a la monotonía: {trm_txt} (TRM = {self.TRM:+.1f} %).')
        parts.append(f'Tiempo de reacción medio por ítem: {self.mean_rt_ms:.0f} ms.')
        parts.append(f'Estilo atencional predominante: {self.attn_style}.')
        return '  '.join(parts)

    # =========================================================================
    # PANTALLA DE RESULTADOS — con gráficas integradas
    # =========================================================================

    def show_results(self):
        self.clear()
        C = self.C
        self._hdr('PLC Professional — Resultados',
                  f'{self.pname}  ·  ID: {self.pid}  ·  '
                  f'{datetime.now().strftime("%d/%m/%Y  %H:%M")}')

        inner, cv_s = self._scroll_frame()
        frame = tk.Frame(inner, bg=C['bg'], padx=35, pady=18)
        frame.pack(fill='both', expand=True)

        # ── A) MÉTRICAS CLAVE ──────────────────────────────────────────────
        s1 = tk.LabelFrame(frame, text='  Métricas Objetivas  ',
                           font=('Georgia', 13, 'bold'),
                           bg=C['white'], fg=C['primary'],
                           padx=18, pady=14, relief='flat', bd=1)
        s1.pack(fill='x', pady=8)

        # Fila 1 — métricas principales
        r1 = tk.Frame(s1, bg=C['white']); r1.pack(fill='x', pady=4)
        for lbl, val, fg, bg in [
            ('TA  Aciertos',  str(self.TA),       C['success'], C['s_bg']),
            ('O  Omisiones',  str(self.O),         C['warn'],    C['w_bg']),
            ('C  Comisiones', str(self.COM),         C['err'],     C['e_bg']),
            ('CP %\nÍndice Concentración', f'{self.CP:.1f}',
             C['primary'], C['a_light']),
            ('CON  Neto',     str(self.CON),       C['secondary'],C['a_light']),
        ]:
            card = tk.Frame(r1, bg=bg, padx=14, pady=10)
            card.pack(side='left', padx=5, expand=True, fill='x')
            tk.Label(card, text=val, font=('Georgia', 24, 'bold'),
                     bg=bg, fg=fg).pack()
            tk.Label(card, text=lbl, font=('Helvetica', 9),
                     bg=bg, fg=C['t_lite'], justify='center').pack()

        # Fila 2 — métricas extendidas
        tk.Frame(s1, bg=C['border'], height=1).pack(fill='x', pady=8)
        r2 = tk.Frame(s1, bg=C['white']); r2.pack(fill='x')
        for lbl, val in [
            ('Velocidad',          f'{self.proc_speed:.0f} estím/min'),
            ('Estabilidad',        f'{self.estabilidad:.0f} %'),
            ('TRM Monotonía',      f'{self.TRM:+.1f} %'),
            ('IVR Vel–Exactitud',  f'{self.IVR:.2f}'),
            ('TR medio',           f'{self.mean_rt_ms:.0f} ms'),
            ('Estilo Atencional',  self.attn_style),
        ]:
            c = tk.Frame(r2, bg=C['a_light'], padx=10, pady=8)
            c.pack(side='left', padx=4, expand=True, fill='x')
            tk.Label(c, text=val, font=('Helvetica', 11, 'bold'),
                     bg=C['a_light'], fg=C['primary']).pack()
            tk.Label(c, text=lbl, font=('Helvetica', 8),
                     bg=C['a_light'], fg=C['t_lite']).pack()

        # ── B) GRÁFICAS (solo si matplotlib disponible) ────────────────────
        if MPL_AVAILABLE:
            self._draw_charts(frame)
        else:
            self._draw_charts_tk(frame)   # fallback Canvas puro

        # ── C) PERFIL COGNITIVO (MLP) ──────────────────────────────────────
        s3 = tk.LabelFrame(frame, text='  Perfil Cognitivo — Indicadores  ',
                           font=('Georgia', 13, 'bold'),
                           bg=C['white'], fg=C['primary'],
                           padx=18, pady=14, relief='flat', bd=1)
        s3.pack(fill='x', pady=8)

        if self.ml_pred and self.ml_pred.get('model_used'):
            pi   = self.ml_pred['profile_info']
            ph   = tk.Frame(s3, bg=C['white']); ph.pack(fill='x', pady=(0,8))
            tk.Label(ph, text=pi['nombre'],
                     font=('Georgia', 16, 'bold'),
                     bg=C['white'], fg=C['primary']).pack(side='left')
            tk.Label(ph,
                     text=f"  ·  Confianza estadística: {self.ml_pred['confidence_percent']}",
                     font=('Helvetica', 11), bg=C['white'],
                     fg=C['t_lite']).pack(side='left', pady=2)

            dc = tk.Frame(s3, bg=C['a_light'], padx=14, pady=10)
            dc.pack(fill='x', pady=4)
            tk.Label(dc, text='Descripción del indicador:',
                     font=('Helvetica', 11, 'bold'),
                     bg=C['a_light'], fg=C['primary']).pack(anchor='w')
            tk.Label(dc, text=pi['desc'], font=('Helvetica', 11),
                     bg=C['a_light'], fg=C['text'],
                     wraplength=880, justify='left').pack(anchor='w', pady=(4,0))

            rf = tk.Frame(s3, bg=C['white']); rf.pack(fill='x', pady=4)
            tk.Label(rf, text='Rasgos observados:',
                     font=('Helvetica', 11, 'bold'),
                     bg=C['white'], fg=C['primary']).pack(anchor='w')
            for r in pi['rasgos']:
                tk.Label(rf, text=f'   •  {r}', font=('Helvetica', 11),
                         bg=C['white'], fg=C['text']).pack(anchor='w')

            # Barra de probabilidades top-4
            probs = sorted(self.ml_pred['all_probs'].items(),
                           key=lambda x: x[1], reverse=True)[:4]
            pf = tk.Frame(s3, bg=C['white']); pf.pack(fill='x', pady=6)
            tk.Label(pf, text='Distribución de indicadores (modelo):',
                     font=('Helvetica', 10, 'bold'),
                     bg=C['white'], fg=C['t_lite']).pack(anchor='w', pady=(0,4))
            pb_row = tk.Frame(pf, bg=C['white']); pb_row.pack(fill='x')
            for pname, prob in probs:
                is_p = (pname == self.ml_pred['predicted_profile'])
                bg = C['accent'] if is_p else C['a_light']
                fg = C['white']  if is_p else C['text']
                pc = tk.Frame(pb_row, bg=bg, padx=10, pady=6)
                pc.pack(side='left', padx=4)
                tk.Label(pc, text=pname.replace('_',' '),
                         font=('Helvetica', 9, 'bold' if is_p else 'normal'),
                         bg=bg, fg=fg).pack()
                tk.Label(pc, text=f'{prob*100:.1f} %',
                         font=('Helvetica', 12, 'bold'),
                         bg=bg, fg=fg).pack()
        else:
            tk.Label(s3,
                     text='Modelo IA no disponible. Las métricas objetivas son válidas para el análisis.',
                     font=('Helvetica', 11),
                     bg=C['white'], fg=C['t_lite'],
                     wraplength=880).pack(anchor='w', pady=8)

        # ── D) NARRATIVA TÉCNICA ───────────────────────────────────────────
        s4 = tk.LabelFrame(frame, text='  Descripción del Rendimiento Cognitivo  ',
                           font=('Georgia', 13, 'bold'),
                           bg=C['white'], fg=C['primary'],
                           padx=18, pady=14, relief='flat', bd=1)
        s4.pack(fill='x', pady=8)
        tk.Label(s4, text=self._narrative(),
                 font=('Helvetica', 11), bg=C['white'], fg=C['text'],
                 wraplength=920, justify='left').pack(anchor='w')

        # ── BOTONES FINALES ────────────────────────────────────────────────
        bf = tk.Frame(frame, bg=C['bg'], pady=16); bf.pack(fill='x')
        self._btn(bf, '🏠  Menú',          self.show_menu,   style='secondary', side='left')
        self._btn(bf, '🔄  Nueva eval.',   self.show_form,   style='ghost',     side='left')
        if getattr(self,'excel_file',None) and os.path.exists(self.excel_file):
            self._btn(bf, '📊  Abrir Excel', self._open_excel, style='primary', side='right')

        inner.update_idletasks(); cv_s.configure(scrollregion=cv_s.bbox('all'))

    # ─── GRÁFICAS CON MATPLOTLIB ─────────────────────────────────────────────

    def _draw_charts(self, parent):
        C  = self.C
        ld = self.lines_data
        lineas = [l['linea']    for l in ld]
        hits   = [l['aciertos'] for l in ld]
        oms    = [l['omisiones']for l in ld]
        coms   = [l['comisiones']for l in ld]

        chart_card = tk.LabelFrame(parent,
                                   text='  Gráficas de Rendimiento  ',
                                   font=('Georgia', 13, 'bold'),
                                   bg=C['white'], fg=C['primary'],
                                   padx=12, pady=12, relief='flat', bd=1)
        chart_card.pack(fill='x', pady=8)

        fig = Figure(figsize=(13, 8), facecolor='white', dpi=90)
        fig.subplots_adjust(hspace=0.45, wspace=0.35)

        # ── Gráfica 1: Curva de comportamiento (aciertos por línea) ───────
        ax1 = fig.add_subplot(2, 2, 1)
        ax1.plot(lineas, hits, color='#1565C0', linewidth=2.2,
                 marker='o', markersize=5, label='Aciertos')
        ax1.fill_between(lineas, hits, alpha=0.12, color='#1565C0')
        ax1.set_title('Curva de Comportamiento', fontweight='bold', fontsize=11)
        ax1.set_xlabel('Línea'); ax1.set_ylabel('Aciertos')
        ax1.set_xticks(lineas); ax1.grid(alpha=0.3)
        # tendencia
        if len(lineas) > 2:
            z = np.polyfit(lineas, hits, 1)
            p = np.poly1d(z)
            ax1.plot(lineas, p(lineas), '--', color='#E53935',
                     linewidth=1.4, label='Tendencia')
        ax1.legend(fontsize=8)

        # ── Gráfica 2: Estilos de atención (barras) ───────────────────────
        ax2 = fig.add_subplot(2, 2, 2)
        labels2  = ['Aciertos\n(TA)', 'Omisiones\n(O)', 'Comisiones\n(C)',
                    'Índice\nConc. CP%']
        values2  = [self.TA, self.O, self.COM, round(self.CP, 1)]
        colors2  = ['#1565C0', '#E65100', '#B71C1C', '#2E7D32']
        bars = ax2.bar(labels2, values2, color=colors2, width=0.55,
                       edgecolor='white', linewidth=1.2)
        for bar, val in zip(bars, values2):
            ax2.text(bar.get_x() + bar.get_width()/2,
                     bar.get_height() + 0.3, str(val),
                     ha='center', va='bottom', fontsize=9, fontweight='bold')
        ax2.set_title('Métricas de Atención', fontweight='bold', fontsize=11)
        ax2.set_ylabel('Valor'); ax2.grid(axis='y', alpha=0.3)

        # ── Gráfica 3: Omisiones y comisiones por línea (barras apiladas) ─
        ax3 = fig.add_subplot(2, 2, 3)
        x = np.array(lineas)
        ax3.bar(x - 0.2, oms,  0.35, label='Omisiones',  color='#E65100', alpha=0.85)
        ax3.bar(x + 0.2, coms, 0.35, label='Comisiones', color='#B71C1C', alpha=0.85)
        ax3.set_title('Errores por Línea', fontweight='bold', fontsize=11)
        ax3.set_xlabel('Línea'); ax3.set_ylabel('Errores')
        ax3.set_xticks(lineas); ax3.legend(fontsize=8); ax3.grid(axis='y', alpha=0.3)

        # ── Gráfica 4: Curva de Normalidad (campana de Gauss) ─────────────
        ax4 = fig.add_subplot(2, 2, 4)
        # Parámetros normativos generados matemáticamente
        # (se reemplazarán con datos reales en versión futura)
        age_v  = int(self.age) if str(self.age).isdigit() else 25
        if age_v <= 18:   mu, sigma = 68, 14
        elif age_v <= 35: mu, sigma = 75, 12
        elif age_v <= 50: mu, sigma = 70, 13
        else:              mu, sigma = 62, 15

        x_n  = np.linspace(mu - 4*sigma, mu + 4*sigma, 300)
        y_n  = (1/(sigma * np.sqrt(2*np.pi))) * np.exp(-0.5*((x_n-mu)/sigma)**2)
        ax4.plot(x_n, y_n, color='#1565C0', linewidth=2.2, label='Distribución normativa')
        ax4.fill_between(x_n, y_n, alpha=0.10, color='#1565C0')

        # Posición del evaluado (CP ajustado)
        score = self.adj_score
        y_ev  = (1/(sigma*np.sqrt(2*np.pi))) * math.exp(-0.5*((score-mu)/sigma)**2)
        ax4.axvline(x=score, color='#E53935', linewidth=2,
                    linestyle='--', label=f'Evaluado: {score:.1f}')
        ax4.scatter([score], [y_ev], color='#E53935', s=60, zorder=5)

        # Sombreado percentil
        pct = (1 + math.erf((score - mu) / (sigma * math.sqrt(2)))) / 2
        x_fill = x_n[x_n <= score]
        y_fill = (1/(sigma*np.sqrt(2*np.pi)))*np.exp(-0.5*((x_fill-mu)/sigma)**2)
        ax4.fill_between(x_fill, y_fill, alpha=0.25, color='#E53935')
        ax4.set_title(f'Curva de Normalidad  (Percentil ≈ {pct*100:.0f})',
                      fontweight='bold', fontsize=11)
        ax4.set_xlabel('CP % ajustado'); ax4.set_ylabel('Densidad')
        ax4.legend(fontsize=8); ax4.grid(alpha=0.3)
        ax4.text(0.02, 0.92,
                 f'Media normativa: {mu}  |  σ: {sigma}\n(Rango etario: {age_v} años)',
                 transform=ax4.transAxes, fontsize=7, color='#546E7A')

        canvas_fig = FigureCanvasTkAgg(fig, master=chart_card)
        canvas_fig.draw()
        canvas_fig.get_tk_widget().pack(fill='both', expand=True)

    def _draw_charts_tk(self, parent):
        """Fallback sin matplotlib — mini barras Canvas."""
        C  = self.C
        ld = self.lines_data
        card = tk.LabelFrame(parent, text='  Rendimiento por Línea  ',
                             font=('Georgia', 13, 'bold'),
                             bg=C['white'], fg=C['primary'],
                             padx=14, pady=12, relief='flat', bd=1)
        card.pack(fill='x', pady=8)
        max_h = max((l['aciertos'] for l in ld), default=1)
        for l in ld:
            r = tk.Frame(card, bg=C['white']); r.pack(fill='x', pady=1)
            tk.Label(r, text=f"L{l['linea']:02d}",
                     font=('Courier', 9, 'bold'),
                     bg=C['white'], fg=C['t_lite'], width=4).pack(side='left')
            bw = max(int((l['aciertos']/max_h)*300), 2)
            tk.Frame(r, bg=C['accent'], width=bw, height=12).pack(side='left', padx=1)
            if l['omisiones']:
                tk.Frame(r, bg=C['warn'],
                         width=max(int((l['omisiones']/max_h)*80),4),
                         height=12).pack(side='left', padx=1)
            if l['comisiones']:
                tk.Frame(r, bg=C['err'],
                         width=max(int((l['comisiones']/max_h)*80),4),
                         height=12).pack(side='left', padx=1)
            tk.Label(r,
                     text=f" {l['aciertos']}✓ {l['omisiones']}O {l['comisiones']}C {l['tiempo_s']:.1f}s",
                     font=('Courier', 8), bg=C['white'], fg=C['t_lite']
                     ).pack(side='left', padx=4)

    # ─── HISTORIAL ───────────────────────────────────────────────────────────

    def show_history(self):
        self.clear()
        self._hdr('Historial de Evaluaciones')
        C = self.C

        frame = tk.Frame(self.root, bg=C['bg'], padx=50, pady=20)
        frame.pack(expand=True, fill='both')
        try:
            files = sorted([f for f in os.listdir(self.results_dir)
                            if f.endswith('.xlsx')], reverse=True)
        except Exception as e:
            tk.Label(frame, text=str(e), bg=C['bg'], fg=C['err']).pack()
            self._btn(frame, '← Volver', self.show_menu, style='ghost')
            return

        if not files:
            tk.Label(frame, text='No hay evaluaciones guardadas.',
                     font=('Helvetica', 14),
                     bg=C['bg'], fg=C['t_lite']).pack(pady=50)
        else:
            lf = tk.Frame(frame, bg=C['white']); lf.pack(fill='both', expand=True)
            sb = ttk.Scrollbar(lf); sb.pack(side='right', fill='y')
            lb = tk.Listbox(lf, font=('Courier', 11),
                            bg=C['white'], fg=C['text'],
                            selectbackground=C['sel'], selectforeground=C['white'],
                            height=14, yscrollcommand=sb.set, relief='flat')
            lb.pack(fill='both', expand=True); sb.config(command=lb.yview)
            for f in files: lb.insert('end', f)

            def open_sel():
                s = lb.curselection()
                if s:
                    fp = os.path.join(self.results_dir, lb.get(s[0]))
                    try:
                        if sys.platform=='win32': subprocess.Popen(f'explorer "{fp}"',shell=True)
                        elif sys.platform=='darwin': subprocess.Popen(['open',fp])
                        else: subprocess.Popen(['xdg-open',fp])
                    except Exception as e: messagebox.showerror('Error',str(e))

            def del_sel():
                s = lb.curselection()
                if s:
                    fn = lb.get(s[0])
                    if messagebox.askyesno('Confirmar', f'¿Eliminar {fn}?'):
                        try:
                            os.remove(os.path.join(self.results_dir,fn))
                            lb.delete(s[0])
                        except Exception as e: messagebox.showerror('Error',str(e))

            br = tk.Frame(frame, bg=C['bg']); br.pack(pady=12)
            self._btn(br, '📂 Abrir', open_sel, style='primary', side='left')
            self._btn(br, '🗑  Eliminar', del_sel, style='danger', side='left')

        self._btn(frame, '← Menú', self.show_menu, style='ghost')

    # ─── ABRIR EXCEL ─────────────────────────────────────────────────────────

    def _open_excel(self):
        try:
            if sys.platform=='win32':
                subprocess.Popen(f'explorer "{self.excel_file}"', shell=True)
            elif sys.platform=='darwin':
                subprocess.Popen(['open', self.excel_file])
            else:
                subprocess.Popen(['xdg-open', self.excel_file])
        except Exception as e:
            messagebox.showerror('Error', str(e))


# =============================================================================
# MAIN
# =============================================================================

def main():
    root = tk.Tk()
    root.title('PLC Professional — Evaluación Cognitiva')
    try:
        sty = ttk.Style(root); sty.theme_use('clam')
    except Exception:
        pass
    PLCApp(root)
    root.mainloop()


if __name__ == '__main__':
    main()
