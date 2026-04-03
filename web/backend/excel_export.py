"""
excel_export.py — Exportación Excel descriptiva y rigurosamente clínica.
Eliminación de sesgos interpretativos, con parámetros estrictamente biométricos y observacionales.
"""
import os
import io
import math
from datetime import datetime
from typing import List, Dict, Any, Optional

import numpy as np
import pandas as pd

import matplotlib
matplotlib.use('Agg')
from matplotlib.figure import Figure

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter

BASE_DIR   = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
EXPORTS_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'exports')
os.makedirs(EXPORTS_DIR, exist_ok=True)

# ── Estilos compartidos ────────────────────────────────────────────────────────
H_FILL  = PatternFill('solid', fgColor='1A237E')
H_FONT  = Font(color='FFFFFF', bold=True, name='Calibri', size=11)
H_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
W_ALIGN = Alignment(wrap_text=True, vertical='top', horizontal='left')

EDU_FILL = PatternFill('solid', fgColor='E8EAF6')
EDU_FONT = Font(color='1A237E', bold=True, italic=True, size=10)
DSCL_FILL= PatternFill('solid', fgColor='FFF9C4')
DSCL_FONT= Font(color='BF360C', bold=True, size=9)

def _style_hdr(ws, row=1):
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill = H_FILL; cell.font = H_FONT; cell.alignment = H_ALIGN

def _set_widths(ws, widths, start_col=1):
    for i, w in enumerate(widths, start_col):
        ws.column_dimensions[get_column_letter(i)].width = w

def _add_educational_header(ws, title: str, description: str, disclaimer: bool = False):
    """Añade cabecera educativa clínica en filas 1-3. El contenido principal comienza en fila 5."""
    ws.merge_cells('A1:E1')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14, color='1A237E')
    
    ws.merge_cells('A2:H2')
    ws['A2'] = description
    ws['A2'].font = EDU_FONT
    ws['A2'].fill = EDU_FILL
    ws['A2'].alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[2].height = 40
    
    if disclaimer:
        ws.merge_cells('A3:H3')
        ws['A3'] = "NOTA CLÍNICA: Este sistema describe parámetros paramétricos puramente observacionales. No emite diagnósticos interpretativos. La ponderación nosológica recae enteramente bajo criterio del profesional de la salud evaluador."
        ws['A3'].font = DSCL_FONT
        ws['A3'].fill = DSCL_FILL
        ws['A3'].alignment = Alignment(wrap_text=True, vertical='center')
        ws.row_dimensions[3].height = 30

# ── Generación de imagen con las 4 gráficas ────────────────────────────────────
def _charts_png(lines_data: List[Dict], metrics: Dict, age_v: int,
                ml_pred: Optional[Dict]) -> io.BytesIO:
    """Reproduce gráficas del panel de resultados + distribución algorítmica."""
    lineas = [l['linea']     for l in lines_data]
    hits   = [l['aciertos']  for l in lines_data]
    oms    = [l['omisiones'] for l in lines_data]
    coms   = [l['comisiones']for l in lines_data]

    n_rows = 3 if (ml_pred and ml_pred.get('model_used')) else 2
    fig = Figure(figsize=(14, n_rows * 4.5), facecolor='white', dpi=110)
    fig.subplots_adjust(hspace=0.52, wspace=0.36)

    # 1. Curva de aciertos
    ax1 = fig.add_subplot(n_rows, 2, 1)
    ax1.plot(lineas, hits, color='#1565C0', lw=2.2, marker='o', ms=5, label='Aciertos Objetivo')
    ax1.fill_between(lineas, hits, alpha=0.12, color='#1565C0')
    if len(lineas) > 2:
        z = np.polyfit(lineas, hits, 1)
        ax1.plot(lineas, np.poly1d(z)(lineas), '--', color='#E53935', lw=1.4, label='Tendencia Lineal')
    ax1.set_title('Fluctuación Visual de Eficacia', fontweight='bold', fontsize=11)
    ax1.set_xlabel('Línea (Fase Temporal)'); ax1.set_ylabel('Aciertos')
    ax1.set_xticks(lineas); ax1.grid(alpha=0.3); ax1.legend(fontsize=8)

    # 2. Métricas de atención
    ax2 = fig.add_subplot(n_rows, 2, 2)
    labels2 = ['Aciertos\n(TA)', 'Omisiones\n(O)', 'Comisiones\n(C)', 'A. Neto\n(CON)']
    values2 = [metrics['TA'], metrics['O'], metrics['COM'], metrics['CON']]
    colors2 = ['#1565C0', '#E65100', '#B71C1C', '#2E7D32']
    bars = ax2.bar(labels2, values2, color=colors2, width=0.55, edgecolor='white', lw=1.2)
    for bar, val in zip(bars, values2):
        ax2.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.3, str(val),
                 ha='center', va='bottom', fontsize=9, fontweight='bold')
    ax2.set_title('Proporción de Ejecución Bruta', fontweight='bold', fontsize=11)
    ax2.set_ylabel('Volumen'); ax2.grid(axis='y', alpha=0.3)

    # 3. Errores por línea
    ax3 = fig.add_subplot(n_rows, 2, 3)
    x = np.array(lineas)
    ax3.bar(x - 0.2, oms,  0.35, label='Omisión (Fallo latencia)',  color='#E65100', alpha=0.85)
    ax3.bar(x + 0.2, coms, 0.35, label='Comisión (Fallo inhibición)', color='#B71C1C', alpha=0.85)
    ax3.set_title('Distribución de Tasas de Falla por Iteración', fontweight='bold', fontsize=11)
    ax3.set_xlabel('Línea'); ax3.set_ylabel('Cantidad de Falla')
    ax3.set_xticks(lineas); ax3.legend(fontsize=8); ax3.grid(axis='y', alpha=0.3)

    # 4. Curva Normativa
    ax4 = fig.add_subplot(n_rows, 2, 4)
    if age_v <= 18:   mu, sigma = 68, 14
    elif age_v <= 35: mu, sigma = 75, 12
    elif age_v <= 50: mu, sigma = 70, 13
    else:             mu, sigma = 62, 15
    x_n = np.linspace(mu - 4 * sigma, mu + 4 * sigma, 300)
    y_n = (1 / (sigma * np.sqrt(2 * np.pi))) * np.exp(-0.5 * ((x_n - mu) / sigma) ** 2)
    ax4.plot(x_n, y_n, color='#1565C0', lw=2.2, label='Expectativa demográfica campana estándar')
    score = metrics.get('adjScore', metrics.get('CP', 0))
    y_ev  = (1 / (sigma * np.sqrt(2 * np.pi))) * math.exp(-0.5 * ((score - mu) / sigma) ** 2)
    ax4.axvline(x=score, color='#E53935', lw=2, ls='--', label=f'Marca del Evaluado: {score:.1f}')
    ax4.scatter([score], [y_ev], color='#E53935', s=60, zorder=5)
    ax4.set_title(f'Ratio de Eficiencia respecto al Grupo Etario (Edad: {age_v}a)', fontweight='bold', fontsize=11)
    ax4.set_xlabel('Proporción de Eficiencia'); ax4.set_ylabel('Distribución')
    ax4.legend(fontsize=8); ax4.grid(alpha=0.3)

    # 5. Distribución MLP Objetivo
    if ml_pred and ml_pred.get('model_used') and 'all_probs' in ml_pred:
        ax5 = fig.add_subplot(n_rows, 2, (5, 6))
        probs_sorted = sorted(ml_pred['all_probs'].items(), key=lambda x: x[1], reverse=True)
        prof_labels  = [k.replace('_', ' ') for k, _ in probs_sorted]
        prof_vals    = [v * 100 for _, v in probs_sorted]
        pred_key     = ml_pred.get('predicted_profile', '')
        bar_colors   = ['#3949AB' if k.replace(' ', '_') == pred_key else '#C5CAE9' for k, _ in probs_sorted]
        bars5 = ax5.barh(prof_labels[::-1], prof_vals[::-1], color=bar_colors[::-1], edgecolor='white', height=0.6)
        for bar, val in zip(bars5, prof_vals[::-1]):
            ax5.text(val + 0.3, bar.get_y() + bar.get_height() / 2, f'{val:.1f}%', va='center', fontsize=9, fontweight='bold', color='#1A1A2E')
        ax5.set_title(f'Similitud Bayesiana con Patrones Objetivos de Respuesta', fontweight='bold', fontsize=11)
        ax5.set_xlabel('Probabilidad Matemática Algorítmica (%)'); ax5.grid(axis='x', alpha=0.3)
        ax5.set_xlim(0, max(max(prof_vals) * 1.18, 100))

    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=110, bbox_inches='tight', facecolor='white')
    buf.seek(0)
    return buf

# ── Exportación Directa a DataFrames ──────────────────────────────────────────
def save_excel(participant: Dict, lines_data: List[Dict], click_log: List[Dict],
               metrics: Dict, ml_pred: Optional[Dict], narrative: str) -> str:
    ts  = datetime.now().strftime('%Y%m%d_%H%M%S')
    fn  = f"PLC_{participant['id']}_{ts}.xlsx"
    fp  = os.path.join(EXPORTS_DIR, fn)

    with pd.ExcelWriter(fp, engine='openpyxl') as writer:
        
        # ── Hoja 1: Resumen Clínico ───────────────────────────────────────────
        df_demog = pd.DataFrame([
            {"Parámetro": "ID Paciente", "Dato": participant['id']},
            {"Parámetro": "Nombre", "Dato": participant['name']},
            {"Parámetro": "Edad Cronométrica", "Dato": participant['age']},
            {"Parámetro": "Género Registrado", "Dato": participant.get('gender','')},
            {"Parámetro": "Educación Aprobada", "Dato": participant.get('education','')},
            {"Parámetro": "Ocupación Cruda", "Dato": participant.get('occupation','')}
        ])
        
        # Mapeo descriptivo algoritmico
        ml_nombre = ml_pred['profile_info']['nombre'] if (ml_pred and ml_pred.get('model_used')) else 'N/A'
        ml_desc   = ml_pred['profile_info']['desc'] if (ml_pred and ml_pred.get('model_used')) else 'N/A'
        
        df_metricas = pd.DataFrame([
            {"Indicador Cuantitativo": "Acumulado de Aciertos (TA)", "Valor Calculado": round(metrics['TA'], 1)},
            {"Indicador Cuantitativo": "Acumulado de Omisiones (O)", "Valor Calculado": round(metrics['O'], 1)},
            {"Indicador Cuantitativo": "Acumulado de Comisiones (COM)", "Valor Calculado": round(metrics['COM'], 1)},
            {"Indicador Cuantitativo": "Volumen Neto Estimado (CON)", "Valor Calculado": round(metrics['CON'], 1)},
            {"Indicador Cuantitativo": "Tasa Proporcional de Concentración (CP %)", "Valor Calculado": f"{round(metrics['CP'], 2)} %"},
            {"Indicador Cuantitativo": "Velocidad Latente Promedio (Estímulos/min)", "Valor Calculado": round(metrics['procSpeed'], 1)},
            {"Indicador Cuantitativo": "Discrepancia Temporal entre Bloques (TRM %)", "Valor Calculado": round(metrics['TRM'], 2)},
            {"Indicador Cuantitativo": "Tiempos Medios de Clic (Reacción ms)", "Valor Calculado": round(metrics['meanRt'], 1)}
        ])

        df_patrones = pd.DataFrame([
            {"Dominio Observacional Algorítmico": "Perfil Matemático Primario: " + ml_nombre},
            {"Dominio Observacional Algorítmico": "Reconocimiento Paramétrico: " + ml_desc},
            {"Dominio Observacional Algorítmico": "Curvatura Explicada Cronométrica: " + metrics.get('attnDesc', 'N/A')}
        ])

        df_demog.to_excel(writer, sheet_name='01_Resumen_Clinico', index=False, startrow=4, startcol=0)
        df_metricas.to_excel(writer, sheet_name='01_Resumen_Clinico', index=False, startrow=4, startcol=3)
        df_patrones.to_excel(writer, sheet_name='01_Resumen_Clinico', index=False, startrow=15, startcol=0)
        
        ws1 = writer.sheets['01_Resumen_Clinico']
        _add_educational_header(ws1, "Resumen Transversal Objetivo del Desempeño",
                                "Esta hoja presenta un consolidado de ejecución biométrica. Contiene parámetros de demografía, totales numéricos y descripciones generadas a paridad poblacional.", 
                                disclaimer=True)
        _style_hdr(ws1, row=5); _style_hdr(ws1, row=16)
        _set_widths(ws1, [25, 30, 2, 40, 20])
        for row in ws1.iter_rows(min_row=5, max_col=5):
            for cell in row: cell.alignment = W_ALIGN

        # ── Hoja 2: Análisis por Línea ─────────────────────────────────────
        d2 = []
        for l in lines_data:
            d2.append({
                "Nº Línea": l['linea'],
                "Estímulos Blancos": l['targets_total'],
                "Exactitud (Marcar Blanco)": l['aciertos'],
                "Fallo por Exclusión (Omisión)": l['omisiones'],
                "Fallo por Inclusión (Comisión)": l['comisiones'],
                "Latencia Invertida (Segundos)": round(l['tiempo_s'], 2),
                "Proporción Exactitud (%)": round((l['aciertos']/max(l['targets_total'],1))*100, 1)
            })
        df_l = pd.DataFrame(d2)
        df_l.to_excel(writer, sheet_name='02_Analisis_Lineas', index=False, startrow=4)
        
        ws2 = writer.sheets['02_Analisis_Lineas']
        _add_educational_header(ws2, "Disección Iterativa de la Ejecución (Línea por Línea)",
                                "Desglose longitudinal de la instrumentación. Útil para ubicar focos precisos de aparición de fallos por latencia crónica o desgaste temprano. (Targets Totales = Exactitud + Omisión).")
        _style_hdr(ws2, row=5)
        _set_widths(ws2, [12, 18, 25, 25, 25, 25, 25])
        
        # ── Hoja 3: Glosario de Métricas ───────────────────────────────────
        glosario = [
            {"Acrónimo": "TA",  "Terminología Clínica": "Total Aciertos", "Definición Teórica Obj.": "Suma pura de blancos marcados correctamente en el espectro visual."},
            {"Acrónimo": "O",   "Terminología Clínica": "Omisiones", "Definición Teórica Obj.": "Estímulos que correspondían al rasgo diana pero que el evaluado esquivó o no vió en la barrida visomotora."},
            {"Acrónimo": "COM", "Terminología Clínica": "Comisiones", "Definición Teórica Obj.": "Marcaciones falsas positivas. Fallo de la función inhibitoria al seleccionar blancos distractores creyéndolos diana."},
            {"Acrónimo": "CON", "Terminología Clínica": "Acierto Neto", "Definición Teórica Obj.": "Resta del total de aciertos frente al total crudo acumulado de errores cruzados (Omisiones + Comisiones)."},
            {"Acrónimo": "CP%", "Terminología Clínica": "Capacidad de Concentración", "Definición Teórica Obj.": "Porcentaje normalizado de exactitud (CON / Total de Tareas Existentes). Refleja pulcritud en filtrado visual."},
            {"Acrónimo": "TRM", "Terminología Clínica": "Tasa de Variación", "Definición Teórica Obj.": "Desviación estático lineal de exactitud entre las líneas de inicio vs cierre. Disminuciones altas revelan pérdida rápida de rendimiento iterativo."},
            {"Acrónimo": "IVR", "Terminología Clínica": "Índice Velocidad/Acierto", "Definición Teórica Obj.": "Progresión logarítmica para estimar el costo temporal que el usuario gasta involuntariamente en evitar cada comisión."}
        ]
        pd.DataFrame(glosario).to_excel(writer, sheet_name='03_Glosario_Metricas', index=False, startrow=4)
        ws3 = writer.sheets['03_Glosario_Metricas']
        _add_educational_header(ws3, "Glosario Biomédico Estandarizado",
                                "Diccionario de referencia analítica de lectura. Detalla la justificación aritmética y la naturaleza neurológica base para cada cuantificador empleado en los reportes PLC.")
        _style_hdr(ws3, row=5)
        _set_widths(ws3, [15, 25, 95])
        for row in ws3.iter_rows(min_row=6, max_col=3):
            row[2].alignment = W_ALIGN

        # ── Hoja 4: Registro de Eventos ────────────────────────────────────
        df_click = pd.DataFrame(click_log) if click_log else pd.DataFrame(columns=['line','stim_idx','is_target','stim_key','action','elapsed_ms'])
        # Rename columns to clinical observation tone
        if not df_click.empty:
            df_click = df_click.rename(columns={
                'line': 'Línea Ejecutada', 'stim_idx': 'Columna de Ojo', 
                'is_target': 'Blanco Requerido', 'stim_key': 'Código de Ítem',
                'action': 'Acción Emitida', 'elapsed_ms': 'Latencia de Motor (ms)'
            })
        df_click.to_excel(writer, sheet_name='04_Registro_Eventos_CRUDOS', index=False, startrow=4)
        ws4 = writer.sheets['04_Registro_Eventos_CRUDOS']
        _add_educational_header(ws4, "Auditoría Logarítmica de Frecuencia Muscular Visomotriz",
                                "Registro técnico asíncrono. Monitorea cada 'clic' del mouse o pulsación de membrana táctil ejecutada per mili-segundo para documentaciones minuciosas forenses de respuesta física.")
        _style_hdr(ws4, row=5)
        _set_widths(ws4, [18, 18, 20, 18, 20, 25])

        # ── Hoja 5: Datos Gráficas ─────────────────────────────────────────
        if lines_data:
            gd = {'Línea': [l['linea'] for l in lines_data], 'Exactitud': [l['aciertos'] for l in lines_data], 'Omisiones': [l['omisiones'] for l in lines_data], 'Comisiones': [l['comisiones'] for l in lines_data]}
            pd.DataFrame(gd).to_excel(writer, sheet_name='05_Arrays_Para_Graficas', index=False)
            ws5 = writer.sheets['05_Arrays_Para_Graficas']
            _style_hdr(ws5, row=1); _set_widths(ws5, [10,12,12,12])

    # ── Hoja 6: Imágenes Embebidas ─────────────────────────────────────────
    try:
        wb  = openpyxl.load_workbook(fp)
        ws6 = wb.create_sheet('06_Analisis_Visual')
        ws6.sheet_view.showGridLines = False

        _add_educational_header(ws6, "Esquemas Gráficos de Conducta y Expectativas Modulares",
                                "Representaciones en 2D que agrupan los arrays de métricas y la clasificación bayesiana en polígonos asintomáticos, diseñados para visualizar rápidamente desviaciones del grupo base normativo.\n(Haga Scroll hacia abajo para ver el panel).", disclaimer=True)
        # Fix height since it's an image
        ws6.row_dimensions[2].height = 60

        age_v = participant.get('age', 25)
        chart_buf = _charts_png(lines_data, metrics, age_v, ml_pred)
        img = XLImage(chart_buf)
        img.anchor = 'A5'
        ws6.add_image(img)

        wb.save(fp)
    except Exception as e:
        print(f'⚠️ Error exportando XL_Visuales: {e}')

    return fp
